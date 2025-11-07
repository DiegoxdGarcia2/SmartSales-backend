import io
import csv
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from django.template.loader import render_to_string
from django.conf import settings
from decimal import Decimal
import datetime

logger = logging.getLogger(__name__)

# Lazy import de WeasyPrint (para que no falle en Windows durante desarrollo)
# En producción (Linux/Docker) funcionará correctamente
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError) as e:
    WEASYPRINT_AVAILABLE = False
    logger.warning(f"WeasyPrint no disponible (normal en Windows): {e}")


def generate_excel_report(queryset, headers: list, title: str):
    """
    Genera un archivo Excel en memoria a partir de un queryset (.values()) y headers.
    Devuelve un objeto HttpResponse listo para ser enviado.
    
    Args:
        queryset: QuerySet de Django (resultado de .values())
        headers (list): Lista de strings con nombres de columnas
        title (str): Título del reporte para el archivo y la hoja
    
    Returns:
        HttpResponse: Respuesta HTTP con el archivo Excel adjunto
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title[:30].replace('/', '-')  # Limpiar título para hoja

    # --- Estilos ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # --- Título del Reporte ---
    sheet.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    title_cell = sheet['A1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=16, color="4F81BD")
    title_cell.alignment = Alignment(horizontal="center")
    
    # --- Timestamp (Fila 2) ---
    sheet.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    timestamp_cell = sheet['A2']
    timestamp_cell.value = f"Generado el: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    timestamp_cell.font = Font(italic=True, size=10)
    timestamp_cell.alignment = Alignment(horizontal="right")

    # --- Headers de la Tabla (Fila 4) ---
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col_num)
        cell.value = header_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        sheet.column_dimensions[get_column_letter(col_num)].width = 25  # Ancho estándar

    # --- Escribir Datos (desde Fila 5) ---
    row_num = 5
    # El queryset ya viene de .values(), que es una lista de diccionarios
    data_rows = [list(item.values()) for item in queryset]

    for row_data in data_rows:
        for col_num, cell_value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            
            # Formatear tipos de datos comunes
            if isinstance(cell_value, datetime.datetime):
                # Manejar timezone-aware y naive datetimes
                try:
                    if timezone.is_aware(cell_value):
                        cell_value = cell_value.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M')
                    else:
                        cell_value = cell_value.strftime('%Y-%m-%d %H:%M')
                except Exception as e:
                    # Fallback: usar formato ISO
                    cell_value = str(cell_value)
                cell.alignment = Alignment(horizontal="center")
            elif isinstance(cell_value, datetime.date):
                cell_value = cell_value.strftime('%Y-%m-%d')
                cell.alignment = Alignment(horizontal="center")
            elif isinstance(cell_value, Decimal):
                cell_value = float(cell_value)
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(cell_value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif cell_value is None:
                cell_value = ""
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")
            
            cell.value = cell_value
            cell.border = thin_border
        
        row_num += 1
    
    logger.info(f"Excel generado: {title} con {len(data_rows)} filas.")

    # --- Guardar en Buffer de Memoria ---
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    # --- Crear HttpResponse con headers correctos ---
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    timestamp_file = timezone.now().strftime('%Y%m%d_%H%M%S')
    # Limpiar título para nombre de archivo (sin caracteres especiales)
    safe_title = title.replace(' ', '_').replace('/', '-').replace('\\', '-')
    filename = f"{safe_title}_{timestamp_file}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Length'] = output.tell()
    
    return response


def generate_pdf_report(queryset, headers: list, title: str, original_prompt: str):
    """
    Genera un archivo PDF en memoria a partir de un queryset, headers y un template HTML.
    Devuelve un objeto HttpResponse listo para ser enviado.
    
    Args:
        queryset: QuerySet de Django (resultado de .values())
        headers (list): Lista de strings con nombres de columnas
        title (str): Título del reporte
        original_prompt (str): Prompt original del usuario para incluir en el PDF
    
    Returns:
        HttpResponse: Respuesta HTTP con el archivo PDF adjunto
    """
    
    # Verificar si WeasyPrint está disponible (Windows sin GTK fallará)
    if not WEASYPRINT_AVAILABLE:
        logger.error("WeasyPrint no está disponible en este sistema (falta GTK runtime)")
        return HttpResponse(
            "PDF generation is not available on this system. "
            "WeasyPrint requires GTK runtime libraries (available in production/Docker). "
            "Please use Excel format or test PDF generation in production.",
            status=503,  # Service Unavailable
            content_type="text/plain"
        )
    
    # ✅ Google Cloud Run: 2GB RAM - Sin límite de registros
    # Procesamos todos los datos del queryset
    data_list = list(queryset)
    
    logger.info(f"📄 Generando PDF con {len(data_list)} registros")
    
    # Formatear datos para el PDF (convertir datetime, Decimal, etc a strings legibles)
    formatted_data = []
    for row_dict in data_list:
        formatted_row = {}
        for key, value in row_dict.items():
            if isinstance(value, datetime.datetime):
                # Manejar timezone-aware y naive datetimes
                try:
                    if timezone.is_aware(value):
                        formatted_row[key] = value.astimezone(timezone.get_current_timezone()).strftime('%d/%m/%Y %H:%M')
                    else:
                        formatted_row[key] = value.strftime('%d/%m/%Y %H:%M')
                except Exception:
                    formatted_row[key] = str(value)
            elif isinstance(value, datetime.date):
                formatted_row[key] = value.strftime('%d/%m/%Y')
            elif isinstance(value, Decimal):
                formatted_row[key] = f"{value:,.2f}"
            elif isinstance(value, (int, float)):
                formatted_row[key] = f"{value:,}"
            elif value is None:
                formatted_row[key] = "N/A"
            else:
                formatted_row[key] = str(value)[:100]  # Truncar strings largos
        formatted_data.append(formatted_row)

    # Renderizar el template HTML a una string
    context = {
        'title': title,
        'headers': headers,
        'data': formatted_data,
        'timestamp': timezone.now(),
        'original_prompt': original_prompt
    }
    html_string = render_to_string('reports/pdf_template.html', context)
    
    # Generar PDF con WeasyPrint (con optimizaciones de memoria)
    try:
        logger.info("🔄 Iniciando generación de PDF con WeasyPrint...")
        
        # Opciones de optimización para reducir uso de memoria
        from weasyprint import HTML
        pdf_file = HTML(string=html_string, base_url=str(settings.BASE_DIR)).write_pdf(
            optimize_size=('fonts', 'images')  # Comprimir fuentes e imágenes
        )
        
        logger.info(f"✅ PDF generado exitosamente: {title} con {len(data_list)} filas ({len(pdf_file)} bytes)")
    except Exception as e:
        logger.error(f"❌ Error al generar PDF con WeasyPrint: {e}", exc_info=True)
        return HttpResponse(
            f"Error al generar PDF: {e}",
            status=500,
            content_type="text/plain"
        )

    # --- Crear HttpResponse con headers correctos ---
    response = HttpResponse(pdf_file, content_type='application/pdf')
    
    timestamp_file = timezone.now().strftime('%Y%m%d_%H%M%S')
    # Limpiar título para nombre de archivo
    safe_title = title.replace(' ', '_').replace('/', '-').replace('\\', '-')
    filename = f"{safe_title}_{timestamp_file}.pdf"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Length'] = len(pdf_file)
    
    return response


def generate_csv_report(queryset, headers: list, title: str):
    """
    Genera un archivo CSV en memoria a partir de un queryset (.values()) y headers.
    Devuelve un objeto HttpResponse listo para ser enviado.
    
    Args:
        queryset: QuerySet de Django (resultado de .values())
        headers (list): Lista de strings con nombres de columnas
        title (str): Título del reporte (usado en el nombre del archivo)
    
    Returns:
        HttpResponse: Respuesta HTTP con el archivo CSV adjunto
    """
    # Crear buffer de memoria
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Escribir headers
    writer.writerow(headers)
    
    # Convertir queryset a lista de diccionarios
    data_list = list(queryset)
    
    # Escribir datos
    for row_dict in data_list:
        formatted_row = []
        for value in row_dict.values():
            # Formatear tipos de datos comunes
            if isinstance(value, datetime.datetime):
                # Manejar timezone-aware y naive datetimes
                try:
                    if timezone.is_aware(value):
                        formatted_value = value.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        formatted_value = value.strftime('%Y-%m-%d %H:%M:%S')
                except Exception:
                    formatted_value = str(value)
            elif isinstance(value, datetime.date):
                formatted_value = value.strftime('%Y-%m-%d')
            elif isinstance(value, Decimal):
                formatted_value = f"{value:.2f}"
            elif value is None:
                formatted_value = ""
            else:
                formatted_value = str(value)
            
            formatted_row.append(formatted_value)
        
        writer.writerow(formatted_row)
    
    logger.info(f"CSV generado: {title} con {len(data_list)} filas.")
    
    # Crear HttpResponse con headers correctos
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8-sig')
    
    timestamp_file = timezone.now().strftime('%Y%m%d_%H%M%S')
    # Limpiar título para nombre de archivo (sin caracteres especiales)
    safe_title = title.replace(' ', '_').replace('/', '-').replace('\\', '-')
    filename = f"{safe_title}_{timestamp_file}.csv"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
